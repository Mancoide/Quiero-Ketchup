#!/usr/bin/env python3
"""
Legacy reconciliation runner used by the panel.

It keeps the old Python-based flow: parse the uploaded files, perform a simple
amount/date/reference reconciliation, write an Excel workbook, and print the
summary JSON expected by Laravel.

MOTOR: Reconciliación Mejorado 2025-05
- Extrae saldos principales reales
- Matching mejorado de referencias
- Excel con formato exacto de contabilidad
"""

import json
import sys
import re
from datetime import datetime
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

from parse_pdf import parse_bank_statement, parse_amount
from parse_spreadsheet import parse_spreadsheet

# ============================================================
# CONFIGURACION DEL MOTOR
# ============================================================
MOTOR_VERSION = "2025-05-v1"
LOG_ENABLED = True

def log(msg):
    """Print log message to stderr for visibility."""
    if LOG_ENABLED:
        print(f"[MOTOR] {msg}", file=sys.stderr)



def parse_file(path):
    extension = Path(path).suffix.lower()
    
    log(f"Parseando archivo: {Path(path).name} (tipo: {extension})")

    if extension == ".pdf":
        return parse_bank_statement(path)

    if extension in [".csv", ".xlsx", ".xls"]:
        return parse_spreadsheet(path)

    if extension == ".json":
        with open(path, "r", encoding="utf-8") as handle:
            data = json.load(handle)
            return data if isinstance(data, list) else data.get("transactions", [])

    raise ValueError(f"Formato no soportado: {extension}")


def extract_ledger_balance(transactions):
    """
    Extrae el saldo final del libro mayor.
    Usa siempre el último saldo registrado en los movimientos.
    """
    if not transactions:
        return 0.0
    
    # Buscar saldos finales en las transacciones
    balances = [amount(tx.get("balance")) for tx in transactions if tx.get("balance") not in (None, "")]
    
    if balances:
        final_balance = round(balances[-1], 2)
        log(f"Saldo del mayor extraído del último saldo leído: {final_balance:,.2f}")
        return final_balance
    
    # Si no hay saldo explícito, sumar todas las transacciones
    total = round(sum(amount(tx.get("amount", 0)) for tx in transactions), 2)
    log(f"Saldo del mayor calculado (suma de movimientos): {total:,.2f}")
    return total


def extract_ledger_balance_from_pdf(path):
    """Extrae el saldo final del mayor desde líneas de acumulado del PDF."""
    if Path(path).suffix.lower() != ".pdf":
        return None

    try:
        with pdfplumber.open(path) as pdf:
            text = "\n".join(page.extract_text() or "" for page in pdf.pages)
    except Exception as exc:
        log(f"⚠ No se pudo extraer saldo del mayor desde PDF: {exc}")
        return None

    patterns = [
        r"Total\s+Gral\.\s+Acumulado\s+Total\s+[\d.,]+\s+[\d.,]+\s+([\d.,]+)",
        r"Acumulado\s+Total\s+[\d.,]+\s+[\d.,]+\s+([\d.,]+)",
    ]

    for pattern in patterns:
        matches = re.findall(pattern, text, flags=re.IGNORECASE)
        if matches:
            balance = amount(matches[-1])
            if balance:
                log(f"Saldo del mayor extraído del PDF: {balance:,.2f}")
                return round(balance, 2)

    return None


def extract_bank_balance(transactions):
    """
    Extrae el saldo final del extracto bancario.
    """
    if not transactions:
        return 0.0
    
    # Buscar saldos finales
    balances = [amount(tx.get("balance")) for tx in transactions if tx.get("balance") not in (None, "")]
    
    if balances:
        final_balance = round(balances[-1], 2)
        log(f"Saldo del extracto extraído (de columna balance): {final_balance:,.2f}")
        return final_balance
    
    # Si no hay saldo explícito, sumar todas las transacciones
    total = round(sum(amount(tx.get("amount", 0)) for tx in transactions), 2)
    log(f"Saldo del extracto calculado (suma de movimientos): {total:,.2f}")
    return total




def extract_reference_number(text):
    """
    Extrae números de referencia de un texto.
    Busca: NRO, Nro, número de movimiento, SIPAP, NC (Nota de Crédito), etc.
    
    Ejemplos:
    "DEPOSITO NRO.: 839911 ..." → 839911
    "TRANSF SIPAP 556178" → 556178
    "NC612487" → 612487
    "MOVIMIENTO 2509152005364" → 2509152005364
    """
    text = str(text or "").upper()
    
    # Patrones prioritarios
    patterns = [
        r"NRO[.:\s]*(\d+)",           # NRO. 839911, NRO: 839911
        r"MOVIMIENTO[.:\s]*(\d+)",    # MOVIMIENTO 12345
        r"(?:SIPAP|SPI)[.:\s]*(\d+)", # SIPAP 556178
        r"(?:NC|N/C)[.:\s]*(\d+)",    # NC612487
        r"CAJA[.:\s]*(\d+)",          # CAJA 22307
        r"(?:COMPROBANTE|COMP)[.:\s]*(\d+)",  # COMPROBANTE 12345
        r"(?:REFERENCIA|REF)[.:\s]*(\d+)",    # REFERENCIA 12345
        r"(\d{6,})",                  # Cualquier número de 6+ dígitos
    ]
    
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            ref = match.group(1)
            # Retornar el número más largo encontrado (más específico)
            return ref
    
    return ""


def transaction_reference(tx):
    """Extrae la referencia desde reference y description."""
    return extract_reference_number(f'{tx.get("reference", "")} {tx.get("description", "")}')


def score(bank, company):
    """
    Calcula puntuación de similitud entre transacciones bancarias y contables.
    """
    points = 0
    days = None

    # Coincidencia de montos (peso: 60%)
    if round(bank["amount"], 2) == round(company["amount"], 2):
        points += 60
        log(f"  ✓ Monto coincide: {bank['amount']}")

    # Coincidencia de fechas (peso: 25%)
    if bank["date_obj"] and company["date_obj"]:
        days = abs((bank["date_obj"] - company["date_obj"]).days)
        if days == 0:
            points += 25
            log(f"  ✓ Fecha exacta")
        elif days <= 2:
            points += 20
            log(f"  ✓ Fecha dentro de 2 días ({days}d)")
        elif days <= 7:
            points += 10
            log(f"  ⚠ Fecha dentro de 7 días ({days}d)")
        elif days <= 10:
            points += 15
            log(f"  ⚠ Fecha dentro de 10 días ({days}d)")

    # Coincidencia de referencias (peso: 20%)
    bank_ref = transaction_reference(bank)
    company_ref = transaction_reference(company)
    
    if bank_ref and company_ref and bank_ref == company_ref:
        points += 70
        log(f"  ✓ Referencia coincide: {bank_ref}")
    
    # Coincidencia en descripción (peso: 15%)
    bank_text = normalize(bank["reference"] + " " + bank["description"])
    company_text = normalize(company["reference"] + " " + company["description"])
    
    if bank_ref and company_ref and bank_ref in company_text:
        if points < 80:  # No contar dos veces
            points += 10
    elif company_ref and company_ref in bank_text:
        if points < 80:
            points += 10

    if days is not None and days > 20:
        points = min(points, 69)

    return points


def parse_date(value):
    if not value:
        return None

    if isinstance(value, datetime):
        return value

    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue

    return None


def amount(value):
    """Convierte un valor a float, maneja formato paraguayo."""
    # Usar la función parse_amount mejorada del módulo parse_pdf
    return parse_amount(value)


def normalize(text):
    """Normaliza texto para comparación: mayúsculas, espacios únicos."""
    return " ".join(str(text or "").upper().split())


def clean_description(text):
    """Quita montos repetidos de la descripción."""
    cleaned = str(text or "")
    cleaned = re.sub(r"\s+s/Planilla\s+de\s+\d[\d.,]*", " s/Planilla", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\s+\d[\d.,]*[.,]\d+\s*$", "", cleaned)
    return " ".join(cleaned.split())




def movement_kind(tx):
    """Determina si una transacción es crédito o débito."""
    text = normalize(f'{tx.get("type", "")} {tx.get("description", "")} {tx.get("reference", "")}')

    if any(word in text for word in ["CREDIT", "CREDITO", "CRÉDITO", "HABER", "ABONO", "DEPOSITO", "DEPÓSITO", "TRF", "TRANSF", "INTRBN"]):
        return "credit"

    if any(word in text for word in [
        "DEBIT", "DEBITO", "DÉBITO", "DEBE", "CHEQUE", "PAGO", "CARGO", "FALTANTE",
        "ANTIC", "AGUINALDO", "COMISION", "VACACION", "VACACIONES", "SALARIO",
        "SUELDO", "BONIFICACION", "BENEFICIOS", "GUARDERIA",
    ]):
        return "debit"

    return "unknown"


def is_internal_payroll_or_expense(tx):
    """Movimientos internos del mayor que se muestran aparte."""
    text = normalize(f'{tx.get("description", "")} {tx.get("reference", "")}')
    return any(word in text for word in [
        "ANTIC", "AGUINALDO", "COMISION", "VACACION", "VACACIONES",
        "SALARIO", "SUELDO", "BONIFICACION", "BENEFICIOS", "GUARDERIA",
    ])


def is_opening_entry(text):
    """Detecta asientos de apertura que no deben mostrarse ni conciliarse."""
    normalized = normalize(text)
    return (
        "ASIENTO DE APERTURA" in normalized
        or "ASIENDO DE APERTURA" in normalized
        or "ASIENTO APERTURA" in normalized
        or "ASIENDO APERTURA" in normalized
    )


def classify_unmatched(tx, origin):
    """Clasifica transacciones no conciliadas por categoría contable."""
    text = normalize(f'{tx.get("description", "")} {tx.get("reference", "")}')
    kind = movement_kind(tx)

    if origin == "mayor":
        excluded_expense_words = [
            "ANTIC", "COMISION", "VACACION", "SALARIO", "SUELDO",
            "BONIFICACION", "BENEFICIOS",
        ]

        if "NO REG" in text and "BANCO" in text and ("DEBITO" in text or "CHEQUE" in text):
            return "cheques_debitos_no_reg_banco"

        if "NO REG" in text and "BANCO" in text and ("DEPOSITO" in text or "CREDITO" in text):
            return "depositos_creditos_no_reg_banco"

        if "ADELANT" in text:
            return "cheques_adelantados"

        if "PENDIENTE" in text:
            return "cheques_pendientes"

        if kind == "debit":
            if "NO REG" in text and "BANCO" in text:
                return "cheques_debitos_no_reg_banco"
            return None

        if kind == "credit" and amount(tx.get("amount")) > 0 and not any(word in text for word in excluded_expense_words):
            return "depositos_creditos_no_reg_banco"

        return None

    if "FALTANTE" in text:
        return "cheques_debitos_no_contabilizados"

    if kind == "credit" or any(word in text for word in ["TRF.INTRBN", "TRF INTRBN", "SIPAP", "SPN", "NC"]):
        return "depositos_creditos_no_contabilizados"

    if kind == "debit":
        return "cheques_debitos_no_contabilizados"

    return None




def reconciliation_category_totals(bank_only, company_only):
    totals = {
        "cheques_pendientes": 0.0,
        "depositos_creditos_no_reg_banco": 0.0,
        "cheques_adelantados": 0.0,
        "cheques_debitos_no_reg_banco": 0.0,
        "cheques_debitos_no_contabilizados": 0.0,
        "depositos_creditos_no_contabilizados": 0.0,
    }

    for tx in company_only:
        category = classify_unmatched(tx, "mayor")
        if category in totals:
            totals[category] += amount(tx.get("amount"))

    for tx in bank_only:
        category = classify_unmatched(tx, "banco")
        if category in totals:
            totals[category] += amount(tx.get("amount"))

    return {key: round(value, 2) for key, value in totals.items()}


def reconciliation_summary(bank_transactions, company_transactions, bank_only, company_only, ledger_balance, bank_balance):
    """
    Genera el resumen de conciliación usando saldos principales reales.
    """
    category_totals = reconciliation_category_totals(bank_only, company_only)
    
    # Usar saldos extraídos del documento
    log(f"Resumen: Saldo Mayor = {ledger_balance:,.2f}, Saldo Banco = {bank_balance:,.2f}")
    
    saldo_conciliado = round(
        ledger_balance
        + category_totals["cheques_pendientes"]
        - category_totals["depositos_creditos_no_reg_banco"]
        + category_totals["cheques_adelantados"]
        - category_totals["cheques_debitos_no_contabilizados"]
        + category_totals["depositos_creditos_no_contabilizados"],
        2,
    )

    return {
        "ledger_balance": ledger_balance,
        "outstanding_checks": category_totals["cheques_pendientes"],
        "bank_unregistered_credits": category_totals["depositos_creditos_no_reg_banco"],
        "advanced_checks": category_totals["cheques_adelantados"],
        "bank_unregistered_debits": category_totals["cheques_debitos_no_reg_banco"],
        "unbooked_debits": category_totals["cheques_debitos_no_contabilizados"],
        "unbooked_credits": category_totals["depositos_creditos_no_contabilizados"],
        "reconciled_balance": saldo_conciliado,
        "bank_statement_balance": bank_balance,
        "difference": round(bank_balance - saldo_conciliado, 2),
    }




def grouped_unmatched(bank_only, company_only):
    grouped = {
        "cheques_pendientes": [],
        "depositos_creditos_no_reg_banco": [],
        "cheques_adelantados": [],
        "cheques_debitos_no_reg_banco": [],
        "cheques_debitos_no_contabilizados": [],
        "depositos_creditos_no_contabilizados": [],
    }

    for tx in company_only:
        category = classify_unmatched(tx, "mayor")
        if category in grouped:
            grouped[category].append(tx)

    for tx in bank_only:
        category = classify_unmatched(tx, "banco")
        if category in grouped:
            grouped[category].append(tx)

    return grouped


def normalize_transactions(rows, source):
    """Normaliza transacciones extrayendo campos estándar."""
    normalized = []
    previous_balance = None

    for index, row in enumerate(rows, start=1):
        description = clean_description(row.get("description") or row.get("descripcion") or row.get("concepto") or "")
        reference = row.get("reference") or row.get("referencia") or ""

        if is_opening_entry(f"{description} {reference}"):
            continue

        tx_debit = amount(row.get("debit") or row.get("debe"))
        tx_credit = amount(row.get("credit") or row.get("haber"))
        tx_amount = amount(row.get("amount") or row.get("monto")) or tx_debit or tx_credit
        if tx_amount == 0:
            continue

        tx_balance = amount(row.get("balance") or row.get("saldo")) if row.get("balance") is not None or row.get("saldo") is not None else None
        tx_type = normalize(row.get("type") or row.get("tipo") or "")

        if tx_debit > 0:
            tx_type = "DEBITO"
        elif tx_credit > 0:
            tx_type = "CREDITO"

        if source == "BANCO" and tx_balance is not None:
            if previous_balance is not None:
                balance_delta = round(tx_balance - previous_balance, 2)
                if abs(abs(balance_delta) - tx_amount) <= 0.01:
                    if balance_delta < 0:
                        tx_debit = tx_amount
                        tx_credit = 0
                        tx_type = "DEBITO"
                    elif balance_delta > 0:
                        tx_debit = 0
                        tx_credit = tx_amount
                        tx_type = "CREDITO"
            elif "DEB.PAGO" in normalize(description):
                tx_debit = tx_amount
                tx_credit = 0
                tx_type = "DEBITO"

            previous_balance = tx_balance

        if not tx_type or tx_type == "UNKNOWN":
            inferred = movement_kind(
                {
                    "type": "",
                    "description": description,
                    "reference": reference,
                }
            )
            tx_type = "CREDITO" if inferred == "credit" else "DEBITO" if inferred == "debit" else ""
        if not tx_type or tx_type == "UNKNOWN":
            tx_type = "DEBITO"

        if tx_debit == 0 and tx_credit == 0:
            if tx_type in ("CREDITO", "CREDIT"):
                tx_credit = tx_amount
            else:
                tx_debit = tx_amount

        if not reference:
            reference = extract_reference_number(description)

        normalized.append(
            {
                "id": f"{source}-{index}",
                "date": row.get("date") or row.get("fecha"),
                "date_obj": parse_date(row.get("date") or row.get("fecha")),
                "description": description,
                "reference": reference,
                "amount": tx_amount,
                "type": tx_type,
                "balance": tx_balance,
                "debit": tx_debit if tx_debit > 0 else None,
                "credit": tx_credit if tx_credit > 0 else None,
                "matched": False,
                "matched_id": None,
            }
        )

    return normalized




def reconcile(bank_transactions, company_transactions):
    """
    Concilia transacciones bancarias con transacciones contables.
    
    Threshold de conciliación: 75 puntos (automático)
    Posibles coincidencias: 60+ puntos (revisar manualmente)
    """
    matches = []
    possible = []

    for bank in bank_transactions:
        best = None
        best_score = 0

        for company in company_transactions:
            if company["matched"]:
                continue

            current_score = score(bank, company)
            if current_score > best_score:
                best = company
                best_score = current_score

        if best and best_score >= 70:
            log(f"COINCIDENCIA: {bank['id']} → {best['id']} (score: {best_score})")
            bank["matched"] = True
            bank["matched_id"] = best["id"]
            best["matched"] = True
            best["matched_id"] = bank["id"]
            matches.append({"bank_id": bank["id"], "company_id": best["id"], "score": best_score})
        elif best and best_score >= 60:
            log(f"POSIBLE: {bank['id']} ↔ {best['id']} (score: {best_score})")
            possible.append({"bank_id": bank["id"], "company_id": best["id"], "score": best_score})

    bank_only = [tx for tx in bank_transactions if not tx["matched"]]
    company_only = [tx for tx in company_transactions if not tx["matched"]]

    log(f"Conciliación: {len(matches)} coincidencias, {len(possible)} posibles, {len(bank_only)} no conciliadas en banco, {len(company_only)} no conciliadas en mayor")

    return matches, bank_only, company_only, possible




def format_currency(value):
    """Formatea valor como moneda paraguaya."""
    return f"{value:,.0f}" if value else "0"


def write_sheet(ws, title, rows):
    """Escribe una hoja de transacciones con formato."""
    ws.append([title])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([])
    ws.append(["Fecha", "Referencia", "Descripcion", "Debe", "Haber", "Saldo"])

    for cell in ws[3]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")

    for row in rows:
        debit = amount(row.get("debit"))
        credit = amount(row.get("credit"))
        if debit == 0 and credit == 0:
            kind = movement_kind(row)
            debit = row["amount"] if kind == "debit" else 0
            credit = row["amount"] if kind == "credit" else 0
        ws.append(
            [
                row["date"],
                row["reference"],
                row["description"],
                debit,
                credit,
                row.get("balance"),
            ]
        )

    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=3, value="Totales")
    ws.cell(row=total_row, column=4, value=f"=SUM(D4:D{total_row - 1})")
    ws.cell(row=total_row, column=5, value=f"=SUM(E4:E{total_row - 1})")
    for column in range(3, 6):
        ws.cell(row=total_row, column=column).font = Font(bold=True)

    for column in ("A", "B", "C", "D", "E", "F"):
        ws.column_dimensions[column].width = 22
    ws.column_dimensions["C"].width = 58

    for column in ("D", "E", "F"):
        for cell in ws[column]:
            cell.alignment = Alignment(horizontal="right")
            if isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.startswith("=")):
                money_cell(cell)


def money_cell(cell):
    """Aplica formato de moneda a una celda."""
    cell.number_format = '[$-es-PY]#,##0.00'
    cell.alignment = Alignment(horizontal="right")


def write_reconciliation_section(ws, row, sign, title, formula_sign, total, rows):
    """
    Escribe una sección de conciliación (ejemplo: CHEQUES PENDIENTES).
    Mantiene el formato exacto del Excel contable.
    """
    # Fila de encabezado de sección
    ws.cell(row=row, column=1, value=sign)
    ws.cell(row=row, column=2, value=title)
    ws.cell(row=row, column=3, value=f"FORMULA: {formula_sign}")
    ws.cell(row=row, column=5, value=total)

    for column in range(1, 6):
        ws.cell(row=row, column=column).font = Font(bold=True)
    
    money_cell(ws.cell(row=row, column=5))

    row += 1
    
    # Filas de detalle
    for tx in sorted(rows, key=lambda item: str(item.get("date") or "")):
        ws.cell(row=row, column=1, value=tx.get("date"))
        ws.cell(row=row, column=2, value=normalize(tx.get("description") or tx.get("reference")))
        ws.cell(row=row, column=4, value=amount(tx.get("amount")))
        money_cell(ws.cell(row=row, column=4))
        row += 1

    return row + 1


def write_workbook(path, bank_transactions, company_transactions, matches, bank_only, company_only, summary):
    """
    Genera el Excel de conciliación exactamente como lo hace contabilidad.
    
    Formato:
    - Hoja 1: Resumen de conciliación
    - Hoja 2: Transacciones del banco
    - Hoja 3: Transacciones del mayor
    """
    workbook = Workbook()

    # HOJA 1: RESUMEN DE CONCILIACION
    ws = workbook.active
    ws.title = "Resumen conciliacion"
    grouped = grouped_unmatched(bank_only, company_only)

    # Título principal
    ws["B1"] = "CONCILIACION"
    ws["B1"].font = Font(bold=True, size=14)
    
    # Saldo del mayor
    ws["B3"] = "SALDO DEL MAYOR"
    ws["E3"] = summary["ledger_balance"]
    ws["B3"].font = Font(bold=True)
    money_cell(ws["E3"])

    # Espacio
    row = 5
    
    # Secciones de conciliación
    row = write_reconciliation_section(ws, row, "MAS", "CHEQUES PENDIENTES", "SUMA", summary["outstanding_checks"], grouped["cheques_pendientes"])
    row = write_reconciliation_section(ws, row, "MENOS", "DEPOSITOS (CREDITOS) NO REG. X BANCO", "RESTA", summary["bank_unregistered_credits"], grouped["depositos_creditos_no_reg_banco"])
    row = write_reconciliation_section(ws, row, "MAS", "CHEQUES ADELANTADOS", "SUMA", summary["advanced_checks"], grouped["cheques_adelantados"])
    row = write_reconciliation_section(ws, row, "MENOS", "CHEQUES (DEBITOS) NO CONTABILIZADOS", "RESTA", summary["unbooked_debits"], grouped["cheques_debitos_no_contabilizados"])
    row = write_reconciliation_section(ws, row, "MAS", "DEPOSITOS (CREDITOS) NO CONTABILIZADOS", "SUMA", summary["unbooked_credits"], grouped["depositos_creditos_no_contabilizados"])

    # Resumen final
    ws.cell(row=row + 1, column=2, value="SALDO S/ CONCILIACION")
    ws.cell(row=row + 1, column=5, value=summary["reconciled_balance"])
    ws.cell(row=row + 3, column=2, value="SALDO S/ EXTRACTO")
    ws.cell(row=row + 3, column=5, value=summary["bank_statement_balance"])
    ws.cell(row=row + 5, column=4, value="DIFERENCIA")
    ws.cell(row=row + 5, column=5, value=summary["difference"])
    
    for summary_row in [row + 1, row + 3, row + 5]:
        ws.cell(row=summary_row, column=2).font = Font(bold=True)
        ws.cell(row=summary_row, column=4).font = Font(bold=True)
        ws.cell(row=summary_row, column=5).font = Font(bold=True)
        money_cell(ws.cell(row=summary_row, column=5))

    # Ancho de columnas
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 58
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    operational_company = [tx for tx in company_transactions if is_internal_payroll_or_expense(tx)]

    # HOJA 2: TRANSACCIONES BANCO
    write_sheet(workbook.create_sheet("Banco"), "TRANSACCIONES BANCO", bank_transactions)
    
    # HOJA 3: TRANSACCIONES MAYOR COMPLETO
    write_sheet(workbook.create_sheet("Mayor"), "TRANSACCIONES MAYOR", company_transactions)

    # HOJA 4: MOVIMIENTOS INTERNOS DEL MAYOR
    if operational_company:
        write_sheet(workbook.create_sheet("Mayor internos"), "MOVIMIENTOS INTERNOS DEL MAYOR", operational_company)

    # Guardar
    workbook.save(path)
    log(f"Excel generado: {path}")




def main():
    if len(sys.argv) < 4:
        print(json.dumps({"error": "Uso: reconcile_transactions.py <banco> <mayor> <salida.xlsx>"}), file=sys.stderr)
        return 1

    bank_path, company_path, output_path = sys.argv[1:4]

    log(f"MOTOR: {MOTOR_VERSION}")
    log(f"Iniciando conciliación...")
    log(f"Banco: {Path(bank_path).name}")
    log(f"Mayor: {Path(company_path).name}")
    log(f"Salida: {Path(output_path).name}")

    try:
        # PASO 1: Parsear archivos
        log(f"[1/4] Parseando archivos...")
        bank_raw = parse_file(bank_path)
        company_raw = parse_file(company_path)
        
        log(f"DEBUG: Bank raw count: {len(bank_raw) if bank_raw else 0}")
        log(f"DEBUG: Company raw count: {len(company_raw) if company_raw else 0}")
        
        if not bank_raw:
            log(f"✗ ERROR: No se pudieron leer transacciones del banco")
            raise ValueError("No se pudieron leer transacciones del banco")
        if not company_raw:
            log(f"✗ ERROR: No se pudieron leer transacciones del mayor")
            raise ValueError("No se pudieron leer transacciones del mayor")
        
        log(f"✓ Banco: {len(bank_raw)} transacciones")
        log(f"✓ Mayor: {len(company_raw)} transacciones")

        # PASO 2: Extraer saldos principales
        log(f"[2/4] Extrayendo saldos principales...")
        ledger_balance = extract_ledger_balance(company_raw) or extract_ledger_balance_from_pdf(company_path)
        bank_balance = extract_bank_balance(bank_raw)
        
        if ledger_balance == 0:
            log(f"⚠ ADVERTENCIA: Saldo del mayor es 0 - verifique el archivo")
        if bank_balance == 0:
            log(f"⚠ ADVERTENCIA: Saldo del banco es 0 - verifique el archivo")

        # PASO 3: Normalizar transacciones
        log(f"[3/4] Normalizando transacciones...")
        bank_transactions = normalize_transactions(bank_raw, "BANCO")
        company_transactions = normalize_transactions(company_raw, "MAYOR")
        
        log(f"✓ Transacciones normalizadas")

        # PASO 4: Conciliar
        log(f"[4/4] Conciliando transacciones...")
        matches, bank_only, company_only, possible = reconcile(bank_transactions, company_transactions)
        
        # Generar resumen
        summary = reconciliation_summary(
            bank_transactions, 
            company_transactions, 
            bank_only, 
            company_only,
            ledger_balance,
            bank_balance
        )

        # Crear directorio de salida
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        
        # Generar Excel
        log(f"Generando Excel...")
        write_workbook(output_path, bank_transactions, company_transactions, matches, bank_only, company_only, summary)
        
        log(f"✓ Conciliación completada exitosamente")

        # Respuesta para Laravel
        print(
            json.dumps(
                {
                    "total_bank_records": len(bank_transactions),
                    "total_company_records": len(company_transactions),
                    "matched_records": len(matches),
                    "bank_only_records": len(bank_only),
                    "company_only_records": len(company_only),
                    "possible_matches": len(possible),
                    "summary": summary,
                },
                ensure_ascii=False,
            )
        )

        return 0

    except Exception as e:
        error_msg = str(e)
        log(f"✗ ERROR: {error_msg}")
        print(json.dumps({"error": error_msg}), file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
